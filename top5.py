import pandas as pd
import numpy as np
import requests
import matplotlib.pyplot as plt
from surprise import Dataset, Reader, SVD
from surprise.model_selection import train_test_split
from tabulate import tabulate
from rich.console import Console
from rich.table import Table
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XLImage
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity 

ratings_df = pd.read_csv("C:/Users/bhava/Downloads/MovieReSys/data/raw/ratings.csv")
movies_df = pd.read_csv("C:/Users/bhava/Downloads/MovieReSys/data/raw/movies.csv")
ratings_df_sample = ratings_df.sample(10000, random_state=42)
reader = Reader(rating_scale=(0.5, 5.0))
data = Dataset.load_from_df(ratings_df_sample[['userId', 'movieId', 'rating']], reader)
trainset = data.build_full_trainset()
model = SVD()
model.fit(trainset)
target_user_id = 1
all_movie_ids = ratings_df_sample['movieId'].unique()
user_seen = ratings_df_sample[ratings_df_sample['userId'] == target_user_id]['movieId'].tolist()
unseen_movies = [mid for mid in all_movie_ids if mid not in user_seen]
predictions = [model.predict(target_user_id, movie_id) for movie_id in unseen_movies]
top_n = sorted(predictions, key=lambda x: x.est, reverse=True)[:5]
print(f"\n🎯 Top 5 recommendations for user {target_user_id}:\n")

for pred in top_n:
    movie_title = movies_df[movies_df['movieId'] == int(pred.iid)]['title'].values[0]
    print(f"{movie_title} (Predicted Rating: {pred.est:.2f})")
print('\n')

for pred in top_n:
    movie_row = movies_df[movies_df['movieId'] == int(pred.iid)].iloc[0]
    print(f"{movie_row['title']} | {movie_row['genres']} (Predicted Rating: {pred.est:.2f})")
print('\n')

results = []
for pred in top_n:
    movie_row = movies_df[movies_df['movieId'] == int(pred.iid)].iloc[0]
    results.append({
        'title': movie_row['title'],
        'genres': movie_row['genres'],
        'predicted_rating': round(pred.est, 2)
    })
top_n_df = pd.DataFrame(results)
print(top_n_df)
print('\n')

for pred in top_n:
    movie_row = movies_df[movies_df['movieId'] == int(pred.iid)].iloc[0]
    print(f"{movie_row['title']} | {movie_row['genres']} (Predicted Rating: {pred.est:.2f})")
print('\n')
print(tabulate(top_n_df, headers='keys', tablefmt='fancy_grid'))
print('\n')

def genre_emojis(genres):
    mapping = {
        'Action': '🔥',
        'Adventure': '🗺️',
        'Comedy': '😂',
        'Drama': '🎭',
        'Horror': '👻',
        'Romance': '❤️',
        'Sci-Fi': '👽',
        'Animation': '🎨',
        'Fantasy': '🧙',
        'Crime': '🕵️',
        'Thriller': '🔪'
    }
    return ' '.join(mapping.get(g, '') for g in genres.split('|'))
top_n_df['emoji_genres'] = top_n_df['genres'].apply(genre_emojis)
for i, row in top_n_df.iterrows():
    print(f"{i+1}. {row['title']} {row['emoji_genres']} → Predicted Rating: {row['predicted_rating']}/5")
print('\n')

def display_rich_table(df):
    console = Console()
    table = Table(title="Top Movie Recommendations", style="bold magenta")
    table.add_column("Title", style="cyan", no_wrap=True)
    table.add_column("Genres")
    table.add_column("Rating", justify="right", style="green")
    for _, row in df.iterrows():
        table.add_row(row['title'], row['emoji_genres'], str(row['predicted_rating']))
    console.print(table)
display_rich_table(top_n_df)
print('\n')

top_n_df.to_csv("C:/Users/bhava/Downloads/MovieReSys/outputs/top_n_recommendations.csv", index=False)
excel_path = "C:/Users/bhava/Downloads/MovieReSys/outputs/top_n_recommendations_styled.xlsx"
top_n_df.to_excel(excel_path, index=False)
wb = load_workbook(excel_path)
ws = wb.active
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
header_font = Font(bold=True, color="000000")
for col_num, cell in enumerate(ws[1], start=1):
    cell.fill = header_fill
    cell.font = header_font
    col_letter = get_column_letter(col_num)
    ws.column_dimensions[col_letter].width = 30  
ws.freeze_panes = "A2"
wb.save(excel_path)
print(f"Styled Excel saved as: {excel_path}")
print('\n')

rating_col_index = top_n_df.columns.get_loc("predicted_rating") + 1  
rating_col_letter = get_column_letter(rating_col_index)
rating_range = f"{rating_col_letter}2:{rating_col_letter}{ws.max_row}"
color_scale = ColorScaleRule(
    start_type='min', start_color='FF6347',  
    mid_type='percentile', mid_value=50, mid_color='FFFF99',  
    end_type='max', end_color='90EE90'  
)
ws.conditional_formatting.add(rating_range, color_scale)
wb.save(excel_path)
print(f"Excel updated with color formatting: {excel_path}")
print('\n')

top_n_sorted = top_n_df.sort_values(by="predicted_rating", ascending=True)
plt.figure(figsize=(10, 6))
bars = plt.barh(top_n_sorted["title"], top_n_sorted["predicted_rating"], color='skyblue')
for bar in bars:
    plt.text(bar.get_width() + 0.02, bar.get_y() + bar.get_height()/2,
             f'{bar.get_width():.2f}', va='center')
plt.xlabel("Predicted Rating")
plt.title("Top 5 Movie Recommendations 🎬")
plt.tight_layout()
plt.grid(axis='x', linestyle='--', alpha=0.5)
plt.show()
print('\n')
plt.figure(figsize=(10, 6))

bars = plt.barh(top_n_sorted["title"], top_n_sorted["predicted_rating"], color='mediumseagreen')
for bar in bars:
    plt.text(bar.get_width() + 0.02, bar.get_y() + bar.get_height()/2,
             f'{bar.get_width():.2f}', va='center')
plt.xlabel("Predicted Rating")
plt.title("Top 5 Movie Recommendations 🎬")
plt.tight_layout()
plt.grid(axis='x', linestyle='--', alpha=0.5)

chart_path = "C:/Users/bhava/Downloads/MovieReSys/outputs/top_n_chart.png"
plt.savefig(chart_path, dpi=300)
plt.close()
print(f"Chart saved as: {chart_path}")
print('\n')

img = XLImage(chart_path)
ws.add_image(img, f"A{ws.max_row + 3}")
wb.save(excel_path)
print(f"Chart image added to: {excel_path}")
print('\n')

movies_df['genres'] = movies_df['genres'].fillna('')
tfidf = TfidfVectorizer(token_pattern=r'[^|]+') 
tfidf_matrix = tfidf.fit_transform(movies_df['genres'])
print(tfidf_matrix.shape)
print('\n')

def get_user_profile(user_id, ratings_df, tfidf_matrix, movies_df):
    user_ratings = ratings_df[ratings_df['userId'] == user_id]
    user_movies = pd.merge(user_ratings, movies_df, on='movieId')
    movie_indices = user_movies['movieId'].apply(lambda x: movies_df[movies_df['movieId'] == x].index[0])
    user_tfidfs = tfidf_matrix[movie_indices]
    ratings = user_movies['rating'].values.reshape(-1, 1)
    user_profile = (user_tfidfs.multiply(ratings)).sum(axis=0)
    return user_profile
user_profile = get_user_profile(1, ratings_df, tfidf_matrix, movies_df)
print('\n')

user_profile = np.asarray(user_profile).reshape(1, -1) 
cosine_sim = cosine_similarity(np.asarray(user_profile), tfidf_matrix)
similarities = list(enumerate(cosine_sim.flatten()))
similarities = sorted(similarities, key=lambda x: x[1], reverse=True)
seen_movie_ids = ratings_df[ratings_df['userId'] == 1]['movieId'].tolist()
recommendations = [(movies_df.iloc[i]['title'], score) for i, score in similarities if movies_df.iloc[i]['movieId'] not in seen_movie_ids]
for title, score in recommendations[:5]:
    print(f"{title} (Score: {score:.4f})")
print('\n')

# TMDB_API_KEY = 'e27947a52a97677530fdd9e476e8b17c'
# TMDB_BASE_URL = 'https://api.themoviedb.org/3'
# def get_movie_id(title):
#     search_url = f"{TMDB_BASE_URL}/search/movie"
#     params = {
#         "api_key": TMDB_API_KEY,
#         "query": title
#     }
#     res = requests.get(search_url, params=params)
#     results = res.json().get('results', [])
#     if results:
#         return results[0]['id']
#     return None
# def get_watch_providers(movie_id, country='US'):
#     url = f"{TMDB_BASE_URL}/movie/{movie_id}/watch/providers"
#     params = {"api_key": TMDB_API_KEY}
#     res = requests.get(url, params=params)
#     data = res.json()
#     return data.get('results', {}).get(country, {}).get('flatrate', [])
# movie_titles = [
#     "The Shawshank Redemption",
#     "Pulp Fiction",
#     "The Godfather",
#     "The Lord of the Rings: The Return of the King",
#     "Fight Club"
# ]
# for title in movie_titles:
#     movie_id = get_movie_id(title)
#     if movie_id:
#         providers = get_watch_providers(movie_id)
#         print(f"{title}:")
#         for provider in providers:
#             print(f"  - {provider['provider_name']}")
#         print()
#     else:
#         print(f"{title}: Not found\n")